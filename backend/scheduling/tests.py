import io
from datetime import date
from unittest.mock import patch

import openpyxl
from django.contrib.auth import get_user_model
from django.test import override_settings
from rest_framework.test import APITestCase

from .models import PointOfSale, Promoter, Schedule, ScheduledVisit

User = get_user_model()


# ── Shared helpers ──────────────────────────────────────────────────────────


def _make_user(username="admin"):
    return User.objects.create_user(username=username, password="pass")


def _make_schedule(user, name, period_start, period_end, status="Draft"):
    return Schedule.objects.create(
        name=name,
        period_start=period_start,
        period_end=period_end,
        status=status,
        created_by=user,
    )


def _make_pos(cdb_code="POS001", name="Test POS", city="Athens", is_active=True):
    return PointOfSale.objects.create(
        cdb_code=cdb_code, name=name, city=city, is_active=is_active
    )


def _make_promoter(username="promo1", first_name="Alice", last_name="Smith"):
    return Promoter.objects.create(
        username=username,
        first_name=first_name,
        last_name=last_name,
        programme_type="Permanent",
    )


def _make_visit(schedule, pos, promoter=None, visit_date=None, **kwargs):
    return ScheduledVisit.objects.create(
        schedule=schedule,
        pos=pos,
        promoter=promoter,
        date=visit_date or date(2026, 4, 3),
        start_time=kwargs.get("start_time", "09:00"),
        end_time=kwargs.get("end_time", "11:00"),
        programme_type=kwargs.get("programme_type", "Permanent"),
        week_label=kwargs.get("week_label", "W1"),
        comments=kwargs.get("comments", ""),
    )


def _make_xlsx(rows):
    """Create an in-memory xlsx with the standard import header + given rows."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(
        [
            "Week",
            "Date",
            "Start Time",
            "End Time",
            "CDB Code",
            "POS Name",
            "City",
            "Priority",
            "Promoter",
            "Programme",
            "AI Reasoning",
        ]
    )
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    buf.name = "schedule.xlsx"
    return buf


# ── Schedule list / auth ────────────────────────────────────────────────────


class ScheduleListAuthTest(APITestCase):
    def test_unauthenticated_returns_401(self):
        response = self.client.get("/api/schedules/")
        self.assertEqual(response.status_code, 401)


class ScheduleListTest(APITestCase):
    def setUp(self):
        self.user = _make_user()
        self.client.force_authenticate(user=self.user)

    def test_empty_list(self):
        response = self.client.get("/api/schedules/")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data, [])

    def test_returns_schedule(self):
        _make_schedule(self.user, "April", date(2026, 4, 1), date(2026, 4, 30))
        response = self.client.get("/api/schedules/")
        self.assertEqual(len(response.data), 1)
        self.assertEqual(response.data[0]["name"], "April")

    def test_ordered_newest_period_first(self):
        _make_schedule(self.user, "March", date(2026, 3, 1), date(2026, 3, 31))
        _make_schedule(self.user, "April", date(2026, 4, 1), date(2026, 4, 30))
        response = self.client.get("/api/schedules/")
        self.assertEqual(response.data[0]["name"], "April")
        self.assertEqual(response.data[1]["name"], "March")

    def test_serializes_expected_fields(self):
        _make_schedule(self.user, "April", date(2026, 4, 1), date(2026, 4, 30))
        data = self.client.get("/api/schedules/").data[0]
        for field in [
            "id",
            "name",
            "period_start",
            "period_end",
            "status",
            "score",
            "created_by",
            "created_at",
            "pos_count",
            "promoter_count",
        ]:
            self.assertIn(field, data)
        self.assertEqual(data["created_by"], "admin")
        self.assertEqual(data["status"], "Draft")

    def test_score_null_by_default(self):
        _make_schedule(self.user, "April", date(2026, 4, 1), date(2026, 4, 30))
        data = self.client.get("/api/schedules/").data[0]
        self.assertIsNone(data["score"])

    def test_pos_count_and_promoter_count(self):
        schedule = _make_schedule(
            self.user, "April", date(2026, 4, 1), date(2026, 4, 30)
        )
        pos = _make_pos()
        promoter = _make_promoter()
        schedule.included_pos.add(pos)
        schedule.included_promoters.add(promoter)
        data = self.client.get("/api/schedules/").data[0]
        self.assertEqual(data["pos_count"], 1)
        self.assertEqual(data["promoter_count"], 1)


class ScheduleListFilterTest(APITestCase):
    def setUp(self):
        self.user = _make_user()
        self.client.force_authenticate(user=self.user)
        _make_schedule(
            self.user, "Draft S", date(2026, 4, 1), date(2026, 4, 30), "Draft"
        )
        _make_schedule(
            self.user, "Published S", date(2026, 5, 1), date(2026, 5, 31), "Published"
        )
        _make_schedule(
            self.user, "Archived S", date(2026, 2, 1), date(2026, 2, 28), "Archived"
        )

    def test_no_filter_returns_all(self):
        response = self.client.get("/api/schedules/")
        self.assertEqual(len(response.data), 3)

    def test_filter_by_draft(self):
        response = self.client.get("/api/schedules/?status=Draft")
        self.assertEqual(len(response.data), 1)
        self.assertEqual(response.data[0]["name"], "Draft S")

    def test_filter_by_published(self):
        response = self.client.get("/api/schedules/?status=Published")
        self.assertEqual(len(response.data), 1)
        self.assertEqual(response.data[0]["status"], "Published")

    def test_filter_by_archived(self):
        response = self.client.get("/api/schedules/?status=Archived")
        self.assertEqual(len(response.data), 1)
        self.assertEqual(response.data[0]["status"], "Archived")

    def test_unknown_status_returns_empty(self):
        response = self.client.get("/api/schedules/?status=Bogus")
        self.assertEqual(len(response.data), 0)


# ── Schedule create ─────────────────────────────────────────────────────────


class ScheduleCreateAuthTest(APITestCase):
    def test_unauthenticated_returns_401(self):
        response = self.client.post("/api/schedules/", {})
        self.assertEqual(response.status_code, 401)


class ScheduleCreateTest(APITestCase):
    def setUp(self):
        self.user = _make_user()
        self.client.force_authenticate(user=self.user)
        self.pos1 = _make_pos("POS001", "POS One")
        self.pos2 = _make_pos("POS002", "POS Two")
        self.promoter1 = _make_promoter("p1", "Alice", "Smith")
        self.promoter2 = _make_promoter("p2", "Bob", "Jones")

    def _payload(self, **overrides):
        data = {
            "name": "April 2026",
            "period_start": "2026-04-01",
            "period_end": "2026-04-30",
            "included_pos": [self.pos1.pk, self.pos2.pk],
            "included_promoters": [self.promoter1.pk],
        }
        data.update(overrides)
        return data

    def test_valid_create_returns_201(self):
        response = self.client.post("/api/schedules/", self._payload(), format="json")
        self.assertEqual(response.status_code, 201)

    def test_creates_with_draft_status(self):
        self.client.post("/api/schedules/", self._payload(), format="json")
        self.assertEqual(Schedule.objects.get().status, "Draft")

    def test_creates_with_current_user(self):
        self.client.post("/api/schedules/", self._payload(), format="json")
        self.assertEqual(Schedule.objects.get().created_by, self.user)

    def test_response_has_list_fields(self):
        response = self.client.post("/api/schedules/", self._payload(), format="json")
        for field in ["id", "name", "status", "pos_count", "promoter_count"]:
            self.assertIn(field, response.data)

    def test_pos_and_promoters_stored(self):
        self.client.post("/api/schedules/", self._payload(), format="json")
        schedule = Schedule.objects.get()
        self.assertEqual(schedule.included_pos.count(), 2)
        self.assertEqual(schedule.included_promoters.count(), 1)

    def test_period_end_before_start_returns_400(self):
        response = self.client.post(
            "/api/schedules/",
            self._payload(period_start="2026-04-30", period_end="2026-04-01"),
            format="json",
        )
        self.assertEqual(response.status_code, 400)

    def test_overlapping_period_returns_400(self):
        _make_schedule(self.user, "Existing", date(2026, 4, 1), date(2026, 4, 30))
        response = self.client.post("/api/schedules/", self._payload(), format="json")
        self.assertEqual(response.status_code, 400)

    def test_missing_name_returns_400(self):
        payload = self._payload()
        del payload["name"]
        response = self.client.post("/api/schedules/", payload, format="json")
        self.assertEqual(response.status_code, 400)


# ── POS / Promoter lists ────────────────────────────────────────────────────


class PointOfSaleListTest(APITestCase):
    def setUp(self):
        self.user = _make_user()
        self.client.force_authenticate(user=self.user)

    def test_unauthenticated_returns_401(self):
        self.client.logout()
        self.assertEqual(self.client.get("/api/pos/").status_code, 401)

    def test_returns_active_pos(self):
        _make_pos("A001", "Active POS", is_active=True)
        _make_pos("A002", "Inactive POS", is_active=False)
        response = self.client.get("/api/pos/")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(response.data), 1)
        self.assertEqual(response.data[0]["cdb_code"], "A001")

    def test_serializes_expected_fields(self):
        _make_pos()
        data = self.client.get("/api/pos/").data[0]
        for field in ["id", "cdb_code", "name", "city", "priority"]:
            self.assertIn(field, data)


class PromoterListTest(APITestCase):
    def setUp(self):
        self.user = _make_user()
        self.client.force_authenticate(user=self.user)

    def test_unauthenticated_returns_401(self):
        self.client.logout()
        self.assertEqual(self.client.get("/api/promoters/").status_code, 401)

    def test_returns_active_promoters(self):
        _make_promoter("active_user", "Alice", "Smith")
        Promoter.objects.create(
            username="inactive_user",
            first_name="Bob",
            last_name="Jones",
            programme_type="Permanent",
            is_active=False,
        )
        response = self.client.get("/api/promoters/")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(response.data), 1)
        self.assertEqual(response.data[0]["username"], "active_user")

    def test_serializes_expected_fields(self):
        _make_promoter()
        data = self.client.get("/api/promoters/").data[0]
        for field in [
            "id",
            "username",
            "first_name",
            "last_name",
            "programme_type",
            "team",
        ]:
            self.assertIn(field, data)


# ── Schedule detail ─────────────────────────────────────────────────────────


class ScheduleDetailTest(APITestCase):
    def setUp(self):
        self.user = _make_user()
        self.client.force_authenticate(user=self.user)
        self.schedule = _make_schedule(
            self.user, "April 2026", date(2026, 4, 1), date(2026, 4, 30)
        )

    def test_returns_schedule(self):
        response = self.client.get(f"/api/schedules/{self.schedule.pk}/")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["name"], "April 2026")

    def test_unauthenticated_returns_401(self):
        self.client.logout()
        self.assertEqual(
            self.client.get(f"/api/schedules/{self.schedule.pk}/").status_code, 401
        )

    def test_not_found_returns_404(self):
        self.assertEqual(self.client.get("/api/schedules/99999/").status_code, 404)


# ── Visit list ──────────────────────────────────────────────────────────────


class ScheduleVisitListTest(APITestCase):
    def setUp(self):
        self.user = _make_user()
        self.client.force_authenticate(user=self.user)
        self.schedule = _make_schedule(
            self.user, "April 2026", date(2026, 4, 1), date(2026, 4, 30)
        )
        self.pos = _make_pos()
        self.promoter = _make_promoter()

    def test_empty_list(self):
        response = self.client.get(f"/api/schedules/{self.schedule.pk}/visits/")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data, [])

    def test_returns_visits(self):
        _make_visit(self.schedule, self.pos, self.promoter)
        response = self.client.get(f"/api/schedules/{self.schedule.pk}/visits/")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(response.data), 1)
        self.assertEqual(response.data[0]["pos"]["cdb_code"], "POS001")

    def test_serializes_nested_pos_and_promoter(self):
        _make_visit(self.schedule, self.pos, self.promoter)
        data = self.client.get(f"/api/schedules/{self.schedule.pk}/visits/").data[0]
        self.assertIn("id", data["pos"])
        self.assertIn("name", data["pos"])
        self.assertIn("first_name", data["promoter"])


# ── AI generation ───────────────────────────────────────────────────────────

_MOCK_AI_RESULT = {
    "summary": "Scheduled 2 visits based on peak windows.",
    "score": 840,
    "visits": [
        {
            "pos_id": None,  # filled in setUp
            "promoter_id": None,
            "date": "2026-04-03",
            "start_time": "09:00",
            "end_time": "11:00",
            "reason": "Peak morning window.",
        },
        {
            "pos_id": None,
            "promoter_id": None,
            "date": "2026-04-10",
            "start_time": "15:00",
            "end_time": "17:00",
            "reason": "Afternoon peak.",
        },
    ],
    "usage": {"prompt_tokens": 500, "completion_tokens": 200, "total_tokens": 700},
}


@override_settings(OPENAI_API_KEY="test-key")
class ScheduleGenerateTest(APITestCase):
    def setUp(self):
        self.user = _make_user()
        self.client.force_authenticate(user=self.user)
        self.pos = _make_pos()
        self.promoter = _make_promoter()
        self.schedule = _make_schedule(
            self.user, "April 2026", date(2026, 4, 1), date(2026, 4, 30)
        )
        self.schedule.included_pos.add(self.pos)
        self.schedule.included_promoters.add(self.promoter)

        self.mock_result = {
            **_MOCK_AI_RESULT,
            "visits": [
                {**v, "pos_id": self.pos.pk, "promoter_id": self.promoter.pk}
                for v in _MOCK_AI_RESULT["visits"]
            ],
        }

    def _post(self, payload=None):
        return self.client.post(
            f"/api/schedules/{self.schedule.pk}/generate/",
            payload or {},
            format="json",
        )

    def test_unauthenticated_returns_401(self):
        self.client.logout()
        self.assertEqual(self._post().status_code, 401)

    @patch("scheduling.views.generate_schedule")
    def test_returns_200_with_visits(self, mock_gen):
        mock_gen.return_value = self.mock_result
        response = self._post({"optimization_goal": "sales * 10 + interviews"})
        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(response.data["visits"]), 2)
        self.assertIn("summary", response.data)
        self.assertIn("usage", response.data)

    @patch("scheduling.views.generate_schedule")
    def test_score_returned_in_response(self, mock_gen):
        mock_gen.return_value = self.mock_result
        response = self._post()
        self.assertEqual(response.data["score"], 840)

    @patch("scheduling.views.generate_schedule")
    def test_score_saved_to_schedule(self, mock_gen):
        mock_gen.return_value = self.mock_result
        self._post()
        self.schedule.refresh_from_db()
        self.assertEqual(self.schedule.score, 840)

    @patch("scheduling.views.generate_schedule")
    def test_clears_existing_visits_before_creating(self, mock_gen):
        mock_gen.return_value = self.mock_result
        _make_visit(self.schedule, self.pos)
        self._post()
        self.assertEqual(
            ScheduledVisit.objects.filter(schedule=self.schedule).count(), 2
        )

    @patch("scheduling.views.generate_schedule")
    def test_visits_saved_to_db(self, mock_gen):
        mock_gen.return_value = self.mock_result
        self._post()
        visits = ScheduledVisit.objects.filter(schedule=self.schedule).order_by("date")
        self.assertEqual(visits.count(), 2)
        self.assertEqual(visits[0].date, date(2026, 4, 3))
        self.assertEqual(str(visits[0].start_time)[:5], "09:00")
        self.assertEqual(visits[0].promoter, self.promoter)
        self.assertEqual(visits[0].comments, "Peak morning window.")

    @patch("scheduling.views.generate_schedule")
    def test_week_label_computed_correctly(self, mock_gen):
        mock_gen.return_value = self.mock_result
        self._post()
        visits = ScheduledVisit.objects.filter(schedule=self.schedule).order_by("date")
        self.assertEqual(visits[0].week_label, "W1")
        self.assertEqual(visits[1].week_label, "W2")

    @patch("scheduling.views.generate_schedule")
    def test_unknown_pos_id_skipped_and_reported(self, mock_gen):
        bad_result = {
            **self.mock_result,
            "visits": [
                {
                    "pos_id": 99999,
                    "promoter_id": self.promoter.pk,
                    "date": "2026-04-03",
                    "start_time": "09:00",
                    "end_time": "11:00",
                    "reason": "test",
                }
            ],
        }
        mock_gen.return_value = bad_result
        response = self._post()
        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(response.data["visits"]), 0)
        self.assertEqual(len(response.data["errors"]), 1)

    @patch("scheduling.views.generate_schedule")
    def test_ai_error_returns_502(self, mock_gen):
        mock_gen.side_effect = Exception("OpenAI timeout")
        response = self._post()
        self.assertEqual(response.status_code, 502)
        self.assertIn("error", response.data)

    @override_settings(OPENAI_API_KEY="")
    def test_missing_api_key_returns_503(self):
        response = self._post()
        self.assertEqual(response.status_code, 503)


# ── Publish ─────────────────────────────────────────────────────────────────


class SchedulePublishTest(APITestCase):
    def setUp(self):
        self.user = _make_user()
        self.client.force_authenticate(user=self.user)
        self.schedule = _make_schedule(
            self.user, "April 2026", date(2026, 4, 1), date(2026, 4, 30), "Draft"
        )

    def _post(self, pk=None):
        pk = pk or self.schedule.pk
        return self.client.post(f"/api/schedules/{pk}/publish/")

    def test_unauthenticated_returns_401(self):
        self.client.logout()
        self.assertEqual(self._post().status_code, 401)

    def test_not_found_returns_404(self):
        self.assertEqual(self._post(pk=99999).status_code, 404)

    def test_draft_publish_returns_200(self):
        self.assertEqual(self._post().status_code, 200)

    def test_status_set_to_published(self):
        self._post()
        self.schedule.refresh_from_db()
        self.assertEqual(self.schedule.status, "Published")

    def test_response_contains_published_status(self):
        response = self._post()
        self.assertEqual(response.data["status"], "Published")

    def test_response_serializes_schedule_fields(self):
        response = self._post()
        for field in ["id", "name", "status", "period_start", "period_end"]:
            self.assertIn(field, response.data)

    def test_publishing_already_published_returns_400(self):
        self.schedule.status = "Published"
        self.schedule.save()
        self.assertEqual(self._post().status_code, 400)

    def test_publishing_archived_returns_400(self):
        self.schedule.status = "Archived"
        self.schedule.save()
        self.assertEqual(self._post().status_code, 400)

    def test_400_response_contains_error_message(self):
        self.schedule.status = "Published"
        self.schedule.save()
        response = self._post()
        self.assertIn("error", response.data)


# ── Export ──────────────────────────────────────────────────────────────────


class ScheduleExportTest(APITestCase):
    def setUp(self):
        self.user = _make_user()
        self.client.force_authenticate(user=self.user)
        self.pos = _make_pos("POS001", "Test POS", city="Athens")
        self.promoter = _make_promoter("alice", "Alice", "Smith")
        self.schedule = _make_schedule(
            self.user, "April 2026", date(2026, 4, 1), date(2026, 4, 30)
        )

    def _get(self, pk=None):
        pk = pk or self.schedule.pk
        return self.client.get(f"/api/schedules/{pk}/export/")

    def test_unauthenticated_returns_401(self):
        self.client.logout()
        self.assertEqual(self._get().status_code, 401)

    def test_not_found_returns_404(self):
        self.assertEqual(self._get(pk=99999).status_code, 404)

    def test_returns_xlsx_content_type(self):
        response = self._get()
        self.assertEqual(response.status_code, 200)
        self.assertIn("spreadsheetml", response["Content-Type"])

    def test_content_disposition_includes_schedule_name(self):
        response = self._get()
        self.assertIn("April 2026", response["Content-Disposition"])
        self.assertIn("attachment", response["Content-Disposition"])

    def test_xlsx_header_row(self):
        response = self._get()
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        self.assertIn("CDB Code", headers)
        self.assertIn("POS Name", headers)
        self.assertIn("Promoter", headers)

    def test_xlsx_contains_visit_data(self):
        _make_visit(
            self.schedule,
            self.pos,
            self.promoter,
            comments="Peak window.",
        )
        response = self._get()
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        self.assertEqual(len(rows), 1)
        # CDB Code is column 5 (index 4)
        self.assertEqual(rows[0][4], "POS001")
        # Promoter is column 9 (index 8)
        self.assertIn("Alice", rows[0][8])

    def test_empty_schedule_returns_header_row_only(self):
        response = self._get()
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        self.assertEqual(len(rows), 0)

    def test_visits_ordered_by_date_then_time(self):
        _make_visit(
            self.schedule,
            self.pos,
            visit_date=date(2026, 4, 10),
            start_time="15:00",
            end_time="17:00",
        )
        _make_visit(
            self.schedule,
            self.pos,
            visit_date=date(2026, 4, 3),
            start_time="09:00",
            end_time="11:00",
        )
        response = self._get()
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        # Date column is index 1
        self.assertEqual(rows[0][1], "2026-04-03")
        self.assertEqual(rows[1][1], "2026-04-10")


# ── Import ──────────────────────────────────────────────────────────────────


class ScheduleImportTest(APITestCase):
    def setUp(self):
        self.user = _make_user()
        self.client.force_authenticate(user=self.user)
        self.pos = _make_pos("POS001", "Test POS")
        self.promoter = _make_promoter("alice", "Alice", "Smith")
        self.schedule = _make_schedule(
            self.user, "April 2026", date(2026, 4, 1), date(2026, 4, 30)
        )

    def _post_xlsx(self, rows):
        buf = _make_xlsx(rows)
        return self.client.post(
            f"/api/schedules/{self.schedule.pk}/import/",
            {"file": buf},
            format="multipart",
        )

    def _valid_row(self, **overrides):
        row = [
            "W1",
            "2026-04-03",
            "09:00",
            "11:00",
            "POS001",
            "Test POS",
            "Athens",
            "Strategic",
            "Alice Smith",
            "Permanent",
            "Peak window.",
        ]
        for i, val in overrides.items():
            row[i] = val
        return row

    def test_unauthenticated_returns_401(self):
        self.client.logout()
        response = self.client.post(
            f"/api/schedules/{self.schedule.pk}/import/", {}, format="multipart"
        )
        self.assertEqual(response.status_code, 401)

    def test_not_found_returns_404(self):
        buf = _make_xlsx([self._valid_row()])
        response = self.client.post(
            "/api/schedules/99999/import/", {"file": buf}, format="multipart"
        )
        self.assertEqual(response.status_code, 404)

    def test_no_file_returns_400(self):
        response = self.client.post(
            f"/api/schedules/{self.schedule.pk}/import/", {}, format="multipart"
        )
        self.assertEqual(response.status_code, 400)

    def test_invalid_file_returns_400(self):
        bad_file = io.BytesIO(b"not an xlsx file")
        bad_file.name = "bad.xlsx"
        response = self.client.post(
            f"/api/schedules/{self.schedule.pk}/import/",
            {"file": bad_file},
            format="multipart",
        )
        self.assertEqual(response.status_code, 400)

    def test_valid_import_creates_visit(self):
        response = self._post_xlsx([self._valid_row()])
        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(response.data["visits"]), 1)
        self.assertEqual(
            ScheduledVisit.objects.filter(schedule=self.schedule).count(), 1
        )

    def test_imported_visit_has_correct_fields(self):
        self._post_xlsx([self._valid_row()])
        visit = ScheduledVisit.objects.get(schedule=self.schedule)
        self.assertEqual(visit.pos, self.pos)
        self.assertEqual(visit.date, date(2026, 4, 3))
        self.assertEqual(str(visit.start_time)[:5], "09:00")
        self.assertEqual(str(visit.end_time)[:5], "11:00")
        self.assertEqual(visit.week_label, "W1")

    def test_import_clears_existing_visits(self):
        _make_visit(self.schedule, self.pos, self.promoter)
        self._post_xlsx([self._valid_row()])
        # Only the imported visit remains
        self.assertEqual(
            ScheduledVisit.objects.filter(schedule=self.schedule).count(), 1
        )

    def test_promoter_resolved_by_full_name(self):
        self._post_xlsx([self._valid_row()])
        visit = ScheduledVisit.objects.get(schedule=self.schedule)
        self.assertEqual(visit.promoter, self.promoter)

    def test_unknown_promoter_name_leaves_promoter_null(self):
        # Override promoter column with unknown name
        row = self._valid_row()
        row[8] = "Unknown Person"
        self._post_xlsx([row])
        visit = ScheduledVisit.objects.get(schedule=self.schedule)
        self.assertIsNone(visit.promoter)

    def test_unknown_cdb_code_skipped_with_error(self):
        row = self._valid_row()
        row[4] = "BADCODE"
        response = self._post_xlsx([row])
        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(response.data["visits"]), 0)
        self.assertEqual(len(response.data["errors"]), 1)
        self.assertIn("BADCODE", response.data["errors"][0])

    def test_date_outside_period_skipped_with_error(self):
        row = self._valid_row()
        row[1] = "2026-06-01"  # outside April
        response = self._post_xlsx([row])
        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(response.data["visits"]), 0)
        self.assertEqual(len(response.data["errors"]), 1)

    def test_missing_date_skipped_with_error(self):
        row = self._valid_row()
        row[1] = None
        response = self._post_xlsx([row])
        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(response.data["visits"]), 0)
        self.assertEqual(len(response.data["errors"]), 1)

    def test_missing_cdb_code_skipped_with_error(self):
        row = self._valid_row()
        row[4] = None
        response = self._post_xlsx([row])
        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(response.data["visits"]), 0)

    def test_multiple_rows_partial_success(self):
        good = self._valid_row()
        bad = self._valid_row()
        bad[4] = "BADCODE"
        response = self._post_xlsx([good, bad])
        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(response.data["visits"]), 1)
        self.assertEqual(len(response.data["errors"]), 1)

    def test_week_label_falls_back_to_computed(self):
        row = self._valid_row()
        row[0] = None  # no week_label in file
        self._post_xlsx([row])
        visit = ScheduledVisit.objects.get(schedule=self.schedule)
        # Apr 3 is day 2 of period starting Apr 1 → W1
        self.assertEqual(visit.week_label, "W1")
